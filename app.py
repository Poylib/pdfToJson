import re
from io import BytesIO
from urllib.parse import urlparse, parse_qs
import html
import hashlib
import zipfile

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests


st.set_page_config(page_title="POSCO IP Agent - RAG JSONL 생성기", page_icon="🛡️", layout="wide")

st.title("🛡️ POSCO IP Agent - RAG JSONL 생성기")
st.caption("엑셀 → 메타/원문 수집 → RAG 최적 청크 JSON 자동 생성")


def normalize_header(value: str) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\u00A0", " ")  # non-breaking space 제거


def extract_url_from_cell(cell) -> tuple[str | None, str | None]:
    """셀에서 URL을 추출하고, 추출 방식(source)을 함께 반환한다.
    우선순위: 실제 하이퍼링크 → HYPERLINK 수식 → 본문 내 URL 텍스트
    """
    # 1) 실제 하이퍼링크 속성
    try:
        if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
            return cell.hyperlink.target, "hyperlink"
    except Exception:
        pass

    # 2) HYPERLINK 수식 형태: =HYPERLINK("url","text") 또는 로캘에 따라 ; 사용 가능성
    try:
        if isinstance(cell.value, str) and cell.value.startswith("=") and "HYPERLINK" in cell.value.upper():
            # 첫 번째 인자로 오는 URL 추출 (따옴표로 둘러싸인 부분)
            m = re.search(r'HYPERLINK\(("|\')(.*?)("|\')', cell.value, re.IGNORECASE)
            if m:
                return m.group(2), "formula"
    except Exception:
        pass

    # 3) 일반 텍스트에 URL 포함
    try:
        if isinstance(cell.value, str):
            m = re.search(r"https?://[^\s)\]\"']+", cell.value)
            if m:
                return m.group(0), "text"
    except Exception:
        pass

    return None, None


def _absolutize_urls(html: str, base: str = "https://sd.wips.co.kr/wipslink/") -> str:
    if not html:
        return html
    # ../path -> absolute
    html = re.sub(r'(src|href)=[\"\']\.\./([^\"\']+)[\"\']', rf"\\1=\"{base}\\2\"", html)
    # //img4.wipson.com -> https://img4.wipson.com
    html = re.sub(r'(src|href)=[\"\']//', r"\\1=\"https://", html)
    return html


def _normalize_description_html(raw_html: str) -> str:
    if not raw_html:
        return ""
    html = raw_html
    # 간단 태그 정규화
    replacements = [
        ("<invention-title>", "<h3>"),
        ("</invention-title>", "</h3>"),
        ("<technical-field>", "<div id=\"technical-field\">"),
        ("</technical-field>", "</div>"),
        ("<background-art>", "<div id=\"background-art\">"),
        ("</background-art>", "</div>"),
        ("<summary-of-invention>", "<div id=\"summary-of-invention\">"),
        ("</summary-of-invention>", "</div>"),
        ("<description-of-drawings>", "<div id=\"description-of-drawings\">"),
        ("</description-of-drawings>", "</div>"),
        ("<description-of-embodiments>", "<div id=\"description-of-embodiments\">"),
        ("</description-of-embodiments>", "</div>"),
        ("<citation-list>", "<div id=\"citation-list\">"),
        ("</citation-list>", "</div>"),
        ("<embodiments-example>", "<div id=\"embodiments-example\">"),
        ("</embodiments-example>", "</div>"),
    ]
    for src, dst in replacements:
        html = html.replace(src, dst)
    # p 태그 단순화
    html = re.sub(r"<p\s+[^>]*>", "<p>", html, flags=re.IGNORECASE)
    # 이미지/링크 절대경로화
    html = _absolutize_urls(html)
    return html


def fetch_wips_description(target_url: str) -> dict:
    """WIPS 상세보기 URL에서 발명의 설명(DS) 전체를 언어별로 수집한다.
    반환: { 'doc': {...meta}, 'descriptions': {langCd: html}, 'order': [lang list] }
    """
    session = requests.Session()
    session.verify = False
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
        "Accept-Language": "ko,en;q=0.9",
    }
    r = session.get(target_url, headers=headers, timeout=20)
    r.raise_for_status()

    # skey 우선 URL 쿼리에서
    parsed = urlparse(target_url)
    qs = parse_qs(parsed.query)
    skey = qs.get("skey", [None])[0]
    # ctry 및 폼 hidden 값 보완
    html = r.text
    if not skey:
        m = re.search(r'id="skey"\s+value=\"(\d+)\"', html)
        if m:
            skey = m.group(1)
    m_ctry = re.search(r'id=\"ctry\"\s+value=\"([A-Z]{2})\"', html)
    ctry = m_ctry.group(1) if m_ctry else "WO"
    if not skey:
        raise RuntimeError("skey를 페이지에서 찾지 못했습니다.")

    post_url = "https://sd.wips.co.kr/wipslink/doc/docContJson.wips"
    data = {
        "skey": skey,
        "tabGb": "DS",
        "isAbEnable": "true",
        "isClEnable": "true",
        "isDsEnable": "true",
        "isAdEnable": "false",
        "isPsEnable": "false",
        "isJdEnable": "false",
        "isFtEnable": "false",
        "devDocType": "DS",
        "devDocCtry": ctry,
    }
    ajax_headers = {
        **headers,
        "Referer": target_url,
        "X-Requested-With": "XMLHttpRequest",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8; metatype=json",
    }
    j = session.post(post_url, headers=ajax_headers, data=data, timeout=30)
    j.raise_for_status()
    payload = j.json()

    # ---------- 메타 보강: 국가/문헌번호 ----------
    doc = payload.get("docPageCmmRsltVO") or {}
    cfg = payload.get("docPageConfigVO") or {}

    def infer_ctry_from_url(path: str) -> str | None:
        if "dkrdshtm" in path:
            return "KR"
        if "dusdshtm" in path:
            return "US"
        if "dwodshtm" in path:
            return "WO"
        if "djpdshtm" in path:
            return "JP"
        if "depdshtm" in path:
            return "EP"
        if "dcndshtm" in path:
            return "CN"
        return None

    meta_ctry = doc.get("ctry") or cfg.get("devDocCtry") or ctry or infer_ctry_from_url(parsed.path) or ""
    nation_code = doc.get("nationCode") or meta_ctry
    composed = " ".join(x for x in [nation_code, doc.get("mngNum"), doc.get("docKind")] if x)
    nation_text = doc.get("nationCodetext") or composed.strip()
    if not nation_text:
        m2 = re.search(r'<p\s+class=\"nation_codetext\">\s*([^<]+)\s*</p>', html)
        if m2:
            nation_text = m2.group(1).strip()

    # ---------- 본문/언어 수집 ----------
    descs = {}
    order = []

    def guess_lang(raw_html: str) -> str:
        if not raw_html:
            return "US"
        kor = len(re.findall(r"[\uac00-\ud7af]", raw_html))
        jpn = len(re.findall(r"[\u3040-\u30ff]", raw_html))
        han = len(re.findall(r"[\u4e00-\u9fff]", raw_html))
        lat = len(re.findall(r"[A-Za-z]", raw_html))
        total = max(kor + jpn + han + lat, 1)
        if kor / total > 0.2:
            return "KR"
        if jpn / total > 0.2:
            return "JP"
        if han / total > 0.4 and kor == 0 and jpn == 0:
            return "CN"
        return "US"

    # 언어별 원문/번역을 분리 저장
    descs_by_lang: dict[str, dict[str, list[str]]] = {}

    def add_desc(lang: str, kind: str, html_content: str):
        bucket = descs_by_lang.setdefault(lang, {"origin": [], "translation": []})
        bucket[kind].append(html_content)

    # 원문(descList)
    for item in (payload.get("descList") or []):
        raw_lang = (item.get("langCd") or "").upper()
        html_part_raw = item.get("dtlDesc") or ""
        html_part = _normalize_description_html(html_part_raw)
        html_part = _absolutize_urls(html_part)
        lang = (raw_lang or guess_lang(html_part_raw) or "UNK").upper()
        add_desc(lang, "origin", html_part)
        if lang not in descs:
            descs[lang] = html_part
            order.append(lang)

    # 번역(descTrnsList)
    for item in (payload.get("descTrnsList") or []):
        raw_lang = (item.get("langCd") or "").upper()
        html_part_raw = item.get("dtlDesc") or ""
        html_part = _normalize_description_html(html_part_raw)
        html_part = _absolutize_urls(html_part)
        lang = (raw_lang or guess_lang(html_part_raw) or "TRNS").upper()
        add_desc(lang, "translation", html_part)
        if lang not in descs and lang not in order:
            descs[lang] = html_part
            order.append(lang)

    # 메타 보강 결과 반영
    if doc is None:
        doc = {}
    doc = {**doc}
    if meta_ctry:
        doc.setdefault("ctry", meta_ctry)
    if nation_text:
        doc.setdefault("nationCodetext", nation_text)
    if cfg.get("devDocCtry"):
        doc.setdefault("devDocCtry", cfg.get("devDocCtry"))

    return {"doc": doc, "descriptions": descs, "order": order, "descriptions_by_lang": descs_by_lang}


def _strip_tags(html: str) -> str:
    if not html:
        return ""
    # script/style 제거
    html = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.IGNORECASE)
    html = re.sub(r"<style[\s\S]*?</style>", " ", html, flags=re.IGNORECASE)
    # 줄바꿈 유도 태그를 개행으로 변환
    html = re.sub(r"<(br|p|li|tr|div)(\s+[^>]*)?>", "\n", html, flags=re.IGNORECASE)
    # 기타 태그 제거
    text = re.sub(r"<[^>]+>", " ", html)
    # HTML 엔티티 간단 치환
    text = text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    # 공백 정리
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def _select_preferred_lang_text(descriptions_by_lang: dict) -> tuple[str, str]:
    """언어 우선순위에 따라 텍스트와 언어코드를 선택한다.
    우선순위: KR(한국어) → EN/US(영어) → 그 외 첫 번째 언어
    반환: (plain_text, selected_lang)
    """
    if not descriptions_by_lang:
        return "", ""

    # KR 우선, 다음 EN/US, 이후 임의 첫 언어
    preferred_langs = ["KR", "EN", "US"]
    for lang in preferred_langs:
        if lang in descriptions_by_lang:
            bucket = descriptions_by_lang.get(lang, {"origin": [], "translation": []})
            htmls = (bucket.get("origin", []) or []) + (bucket.get("translation", []) or [])
            text = _strip_tags("\n".join(htmls))
            return text, lang

    # 위 우선순위에 없으면, 첫 번째 키 사용
    first_lang = next(iter(descriptions_by_lang.keys()))
    bucket = descriptions_by_lang.get(first_lang, {"origin": [], "translation": []})
    htmls = (bucket.get("origin", []) or []) + (bucket.get("translation", []) or [])
    text = _strip_tags("\n".join(htmls))
    return text, first_lang


def _decode_entities_and_normalize(text: str) -> str:
    if not text:
        return ""
    # 숫자/이름 엔티티 2회까지 디코딩 (이중 이스케이프 방지)
    out = html.unescape(text)
    out = html.unescape(out)
    # 줄바꿈 정리 및 공백 정규화
    out = out.replace("\r\n", "\n").replace("\r", "\n")
    # 3개 이상 연속 개행 → 2개로 축소
    out = re.sub(r"\n{3,}", "\n\n", out)
    # 탭을 공백으로 치환 후 다중 공백 축소(개행은 유지)
    out = out.replace("\t", " ")
    out = re.sub(r"[ \f\v]{2,}", " ", out)
    return out.strip()


def _extract_section_hint(text: str) -> str | None:
    if not text:
        return None
    nums = [int(m) for m in re.findall(r"\[(\d{4})\]", text)]
    if not nums:
        return None
    return f"[{min(nums):04d}]-[{max(nums):04d}]"


def _split_by_paragraph_numbers(text: str) -> list[tuple[str, str]]:
    """특허 텍스트를 문단 번호 [0001], [0002] 등으로 분리한다.
    반환: [(paragraph_number, paragraph_text), ...]
    """
    if not text:
        return []

    # [0001], [0002] 형태의 문단 번호로 분리
    pattern = r'(\[\d{4}\])'
    parts = re.split(pattern, text)

    paragraphs = []
    current_num = None

    for i, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue

        # 문단 번호인지 확인
        if re.match(r'^\[\d{4}\]$', part):
            current_num = part
        elif current_num:
            # 문단 번호 다음에 오는 내용
            paragraphs.append((current_num, part))
            current_num = None
        else:
            # 문단 번호 없이 시작하는 텍스트 (제목 등)
            paragraphs.append(("", part))

    return paragraphs


def _chunk_text_with_overlap(text: str, target_chars: int = 4000, overlap_chars: int = 400) -> list[str]:
    """특허 텍스트를 문단 번호 기준으로 의미 단위로 청크 분할한다."""
    if not text:
        return []
    if len(text) <= target_chars:
        return [text]

    # 문단 번호로 분리
    paragraphs = _split_by_paragraph_numbers(text)

    if not paragraphs:
        # 문단 번호가 없으면 기존 방식으로 폴백
        return _chunk_text_simple(text, target_chars, overlap_chars)

    chunks: list[str] = []
    current_chunk: list[str] = []
    current_size = 0

    for para_num, para_text in paragraphs:
        # 문단 번호 + 텍스트
        para_full = f"{para_num} {para_text}".strip() if para_num else para_text
        para_len = len(para_full)

        # 현재 청크에 추가했을 때 크기 확인
        if current_size + para_len > target_chars and current_chunk:
            # 청크 완성
            chunks.append("\n".join(current_chunk))

            # 오버랩을 위해 마지막 몇 개 문단 유지
            overlap_paras = []
            overlap_size = 0
            for p in reversed(current_chunk):
                if overlap_size + len(p) <= overlap_chars:
                    overlap_paras.insert(0, p)
                    overlap_size += len(p)
                else:
                    break

            current_chunk = overlap_paras
            current_size = overlap_size

        current_chunk.append(para_full)
        current_size += para_len

    # 마지막 청크 추가
    if current_chunk:
        chunks.append("\n".join(current_chunk))

    return chunks if chunks else [text]


def _chunk_text_simple(text: str, target_chars: int = 4000, overlap_chars: int = 400) -> list[str]:
    """문단 번호 없는 텍스트를 단순 분할 (폴백용)"""
    if not text:
        return []
    if len(text) <= target_chars:
        return [text]
    chunks: list[str] = []
    start = 0
    end = len(text)
    while start < end:
        cut = min(start + target_chars, end)
        # 문단 경계로 살짝 뒤로 이동
        boundary = text.rfind("\n\n", start + int(0.7 * target_chars), cut)
        if boundary != -1 and boundary > start:
            cut = boundary
        chunk = text[start:cut].strip()
        if chunk:
            chunks.append(chunk)
        if cut >= end:
            break
        start = max(cut - overlap_chars, start + 1)
    return chunks


def _build_chunk_records_for_doc(doc_id: str, base_metadata: dict, raw_text: str, lang: str) -> list[dict]:
    clean_text = _decode_entities_and_normalize(raw_text)
    # 제목 접두사
    title = (base_metadata or {}).get("title") or ""
    prefix = (title + "\n\n") if title else ""
    # 청크화 (문단 번호 기준)
    chunks = _chunk_text_with_overlap(clean_text, target_chars=4000, overlap_chars=400)
    total = len(chunks)
    out: list[dict] = []
    for i, c in enumerate(chunks, start=1):
        text_with_title = prefix + c if prefix else c

        # 문단 번호 범위 추출 ([0001]-[0010] 형식)
        section_hint = _extract_section_hint(c)

        rec = {
            "id": f"{doc_id}-chunk-{i:03d}of{total:03d}",
            "doc_id": doc_id,
            "text": text_with_title,
            "metadata": {
                **base_metadata,
                "lang": lang,
                "chunk_index": i,
                "chunk_total": total,
                "paragraph_range": section_hint or "",  # 더 명확한 이름
            },
        }
        # 선택: 텍스트 해시
        try:
            rec["text_hash"] = hashlib.sha256(text_with_title.encode("utf-8")).hexdigest()
        except Exception:
            pass
        out.append(rec)
    # 빈 텍스트라면 최소 1건 생성 방지
    return out

def _extract_pairs_from_li_html(li_html: str) -> list[dict]:
    """docSummaryInfo > li:first-child 내부에서 보이는 Key-Value 쌍을 최대한 일반적으로 수집한다.
    다양한 마크업(dl/dt+dd, table th/td, span.title+span.value, 텍스트 콜론 구분)을 지원한다.
    반환 형식: [{"label": str, "value": str}, ...] (순서를 보존하려 노력)
    """
    if not li_html:
        return []
    pairs: list[dict] = []

    def add_pair(k: str, v: str):
        k = _strip_tags(k)
        v = _strip_tags(v)
        if k and v:
            pairs.append({"label": k, "value": v})

    # 1) dl/dt/dd 패턴
    for m in re.findall(r"<dt[^>]*>([\s\S]*?)</dt>\s*<dd[^>]*>([\s\S]*?)</dd>", li_html, flags=re.IGNORECASE):
        add_pair(m[0], m[1])

    # 2) table tr (th/td 또는 td/td)
    for tr in re.findall(r"<tr[^>]*>([\s\S]*?)</tr>", li_html, flags=re.IGNORECASE):
        cells = re.findall(r"<(?:th|td)[^>]*>([\s\S]*?)</(?:th|td)>", tr, flags=re.IGNORECASE)
        if len(cells) >= 2:
            key = cells[0]
            val = " | ".join(cells[1:])
            add_pair(key, val)

    # 3) span.title + span.value 형태(클래스명 유사 매칭)
    for m in re.findall(
        r"<span[^>]*class=\"[^\"]*(?:tit|title|key)[^\"]*\"[^>]*>([\s\S]*?)</span>\s*<span[^>]*class=\"[^\"]*(?:cont|val|value|data)[^\"]*\"[^>]*>([\s\S]*?)</span>",
        li_html,
        flags=re.IGNORECASE,
    ):
        add_pair(m[0], m[1])

    # 4) 콜론/화살표 포함 텍스트 라인
    text_block = _strip_tags(li_html)
    for raw_line in [x.strip() for x in text_block.split("\n") if x.strip()]:
        # 우선순위: -->, ->, →, :, ：
        if "-->" in raw_line:
            left, right = raw_line.split("-->", 1)
            add_pair(left, right)
            continue
        if "->" in raw_line:
            left, right = raw_line.split("->", 1)
            add_pair(left, right)
            continue
        if "→" in raw_line:
            left, right = raw_line.split("→", 1)
            add_pair(left, right)
            continue
        if ":" in raw_line or "：" in raw_line:
            sep = ":" if ":" in raw_line else "："
            left, right = raw_line.split(sep, 1)
            add_pair(left, right)

    # 중복 제거(앞선 순서 우선)
    seen = set()
    deduped = []
    for p in pairs:
        key = (p["label"], p["value"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(p)
    return deduped


def _normalize_date(value: str) -> str:
    if not value:
        return ""
    m = re.search(r"(\d{4})[.\-/](\d{2})[.\-/](\d{2})", value)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return value.strip()


def _slugify_key(text: str) -> str:
    t = re.sub(r"\s+", " ", text).strip().lower()
    t = re.sub(r"[()\[\]{}]+", " ", t)
    t = re.sub(r"[^a-z0-9]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t or "field"


def _map_label_to_key(label: str) -> str | None:
    base = re.sub(r"\s+", "", _strip_tags(label)).lower()
    patterns = [
        (r"상태정보|status", "legal_status"),
        (r"최종처분|결정|등록결정|decision", "decision"),
        (r"등록번호|문헌번호|grantno|docnumber|documentnumber", "doc_number"),
        (r"발행일|공고일|issue|gazette", "issue_date"),
        (r"공개번호|publication", "publication_number"),
        (r"관련특허|related", "related_patents"),
        (r"출원번호|application", "application_number"),
        (r"원문상?출원인", "applicants_original"),
        (r"출원인대표|출원인\s*대표명|출원인\s*대표명화|대표자명", "applicant_representative"),
        (r"출원인|applicant", "applicants"),
        (r"현재권리자대표|권리자대표|현재권리자\s*대표명|현재권리자\s*대표명화", "assignee_representative"),
        (r"현재권리자|권리자|assignee", "assignees"),
        (r"출원히스토리|타임라인|history", "timeline"),
    ]
    for pat, key in patterns:
        if re.search(pat, base):
            return key
    return None


def _split_list_like(text: str) -> list[str]:
    if not text:
        return []
    raw = _strip_tags(text)
    parts = [p.strip() for p in re.split(r"[;,/·ㆍ丨|]+", raw) if p.strip()]
    # 중복 제거 순서보존
    seen = set()
    out = []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def structure_docsummary_pairs(pairs: list[dict]) -> dict:
    """라벨-값 리스트를 표준 키를 갖는 JSON으로 변환한다."""
    fields: dict[str, object] = {}
    by_label: dict[str, str] = {}
    extras: dict[str, str] = {}
    for item in pairs:
        label = _strip_tags(item.get("label", ""))
        value = item.get("value", "")
        if not label:
            continue
        key = _map_label_to_key(label)
        clean_val = _strip_tags(value)

        if key in {"issue_date"}:
            val: object = _normalize_date(clean_val)
        elif key in {"applicants", "assignees", "related_patents"}:
            val = _split_list_like(clean_val)
        else:
            val = clean_val

        if key is None:
            extras[_slugify_key(label)] = val  # 의미를 모르면 extras로 보존
        else:
            if key in fields:
                # 기존 값이 있으면 리스트로 병합
                prev = fields[key]
                if isinstance(prev, list):
                    prev = prev + (val if isinstance(val, list) else [val])
                    fields[key] = [x for i, x in enumerate(prev) if x and x not in prev[:i]]
                else:
                    fields[key] = [prev] + (val if isinstance(val, list) else [val])
            else:
                fields[key] = val

        by_label[label] = clean_val

    return {"fields": fields, "by_label": by_label, "extras": extras}


def _extract_pairs_from_text_block(text: str) -> list[dict]:
    """라벨 토막들이 한 줄에 이어지는 형태를 전체 텍스트에서 안정적으로 분리한다.
    규칙: 알려진 라벨 + (--> | -> | → | : | ：) + 값, 다음 라벨이 나오기 전까지를 값으로 간주.
    """
    if not text:
        return []
    raw = re.sub(r"\u00A0", " ", text)
    raw = re.sub(r"\s+", " ", raw).strip()

    # 알려진 라벨 후보(빈도 순으로 배치)
    labels = [
        "상태정보", "최종처분내용", "등록번호", "공고일", "공개번호", "관련특허", "출원번호",
        "출원인", "원문상 출원인", "출원인 대표명", "출원인 대표명화", "현재권리자", "현재권리자 대표명",
        "현재권리자 대표명화", "출원히스토리", "출원인 대표명칭"
    ]
    # 라벨 정규식(라벨에 공백 허용)
    label_regex = r"(?:" + r"|".join([re.escape(l).replace("\\ ", "\\s*") for l in labels]) + r")"
    delim = r"\s*(?:-->|->|→|:|：)\s*"
    pattern = re.compile(rf"({label_regex}){delim}")

    pairs: list[dict] = []
    matches = list(pattern.finditer(raw))
    for i, m in enumerate(matches):
        start_val = m.end()
        end_val = matches[i + 1].start() if i + 1 < len(matches) else len(raw)
        label = m.group(1)
        value = raw[start_val:end_val].strip()
        if label and value:
            pairs.append({"label": label, "value": value})
    return pairs


def fetch_wips_docsummary_first_li(target_url: str) -> dict:
    """상세보기 페이지에서 ul#docSummaryInfo 의 첫 번째 li 내용을 파싱해 반환한다.
    반환: { 'url': url, 'skey': skey, 'pairs': [...], 'raw_html': str }
    """
    session = requests.Session()
    session.verify = False
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
        "Accept-Language": "ko,en;q=0.9",
    }
    r = session.get(target_url, headers=headers, timeout=20)
    r.raise_for_status()
    html = r.text

    parsed = urlparse(target_url)
    qs = parse_qs(parsed.query)
    skey = qs.get("skey", [None])[0]
    if not skey:
        m = re.search(r'id="skey"\s+value="(\d+)"', html)
        if m:
            skey = m.group(1)

    # docSummaryInfo 영역 추출 → 첫 번째 li
    # 정적 HTML 시도(따옴표/공백 변형 허용)
    m_ul = re.search(r"<ul[^>]*id\s*=\s*[\"']docSummaryInfo[\"'][^>]*>([\s\S]*?)</ul>", html, flags=re.IGNORECASE)
    if not m_ul:
        # 비동기 로드 폴백: docContJson.wips에서 탭별 HTML 검색
        post_url = "https://sd.wips.co.kr/wipslink/doc/docContJson.wips"
        ajax_headers = {
            **headers,
            "Referer": target_url,
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8; metatype=json",
        }

        # 국가 코드 시도(없어도 서버가 채움)
        m_ctry = re.search(r'id=\"ctry\"\s+value=\"([A-Z]{2})\"', html)
        ctry = m_ctry.group(1) if m_ctry else "WO"

        tab_candidates = ["OV", "AD", "DS", "CL", "AB", "PS", "JD", "FT"]

        def pick_first_html_with_summary(obj) -> str | None:
            if obj is None:
                return None
            stack = [obj]
            while stack:
                cur = stack.pop()
                if isinstance(cur, dict):
                    for v in cur.values():
                        stack.append(v)
                elif isinstance(cur, list):
                    for v in cur:
                        stack.append(v)
                elif isinstance(cur, str):
                    if "docSummaryInfo" in cur and "<ul" in cur:
                        return cur
            return None

        ul_html = None
        for tab in tab_candidates:
            try:
                data = {
                    "skey": skey or "",
                    "tabGb": tab,
                    "devDocType": tab,
                    "devDocCtry": ctry,
                    # 여러 섹션 활성화 플래그(서버가 무시해도 무방)
                    "isAbEnable": "true",
                    "isClEnable": "true",
                    "isDsEnable": "true",
                    "isAdEnable": "true",
                    "isPsEnable": "true",
                    "isJdEnable": "true",
                    "isFtEnable": "true",
                }
                j = session.post(post_url, headers=ajax_headers, data=data, timeout=30)
                j.raise_for_status()
                payload = j.json()
                html_candidate = pick_first_html_with_summary(payload)
                if not html_candidate:
                    # JSON 전체 직렬화 후 유니코드 이스케이프를 해제하며 검색
                    import json as _json
                    blob = _json.dumps(payload, ensure_ascii=False)
                    try:
                        blob_unesc = bytes(blob, "utf-8").decode("unicode_escape")
                    except Exception:
                        blob_unesc = blob
                    html_candidate = blob_unesc
                if html_candidate:
                    mu = re.search(r"<ul[^>]*id\s*=\s*[\"']docSummaryInfo[\"'][^>]*>([\s\S]*?)</ul>", html_candidate, flags=re.IGNORECASE)
                    if mu:
                        m_ul = mu
                        ul_html = mu.group(1)
                        break
            except Exception:
                continue

        if m_ul is None:
            raise RuntimeError("docSummaryInfo 영역을 찾을 수 없습니다.")

    if 'ul_html' not in locals() or ul_html is None:
        ul_html = m_ul.group(1)
    m_li = re.search(r'<li[^>]*>([\s\S]*?)</li>', ul_html, flags=re.IGNORECASE)
    if not m_li:
        raise RuntimeError("docSummaryInfo 내 첫 번째 li를 찾을 수 없습니다.")
    li_html = m_li.group(1)

    pairs = _extract_pairs_from_li_html(li_html)
    if not pairs:
        # 태그 제거 텍스트 전체에서 라벨 반복 패턴으로 재시도
        pairs = _extract_pairs_from_text_block(_strip_tags(li_html))
    structured = structure_docsummary_pairs(pairs)
    return {"url": target_url, "skey": skey, "pairs": pairs, "structured": structured, "raw_html": li_html}


def _ensure_playwright_ready() -> tuple[bool, str | None]:
    """Playwright 사용 가능 여부를 확인한다. 미설치 시 안내 메시지 반환."""
    try:
        from playwright.sync_api import sync_playwright  # noqa: F401
    except Exception as e:
        return False, f"Playwright 미설치: {e}. 터미널에서 'pip install -r requirements.txt' 후 'python -m playwright install chromium' 실행 필요."
    return True, None


def fetch_docsummary_with_browser(target_url: str, wait_ms: int = 8000) -> dict:
    """헤드리스 브라우저로 페이지를 렌더링해 docSummaryInfo의 첫 li를 수집한다."""
    ok, msg = _ensure_playwright_ready()
    if not ok:
        raise RuntimeError(msg or "Playwright not ready")

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36")
        page = context.new_page()
        page.set_default_timeout(max(wait_ms, 3000))
        page.goto(target_url, wait_until="domcontentloaded")
        # docSummaryInfo가 렌더될 때까지 대기(탭 전환 지연 고려하여 여유롭게)
        try:
            page.wait_for_selector("ul#docSummaryInfo li", state="attached")
        except Exception:
            # 일부 페이지는 탭 전환 필요 → 탭 컨테이너 후보 클릭 시도
            try:
                # '문헌전체' 또는 '개요' 탭 버튼을 클릭해보는 시도
                selectors = ["button[aria-controls='OV']", "#ov", "text=개요", "text=문헌전체"]
                for sel in selectors:
                    try:
                        page.locator(sel).first.click(timeout=1000)
                    except Exception:
                        pass
                page.wait_for_selector("ul#docSummaryInfo li", state="attached")
            except Exception:
                html_snap = page.content()
                browser.close()
                raise RuntimeError("렌더링된 DOM에서 docSummaryInfo를 찾지 못했습니다.")

        # 원시 HTML 스냅샷
        li_html = page.eval_on_selector("ul#docSummaryInfo li:nth-of-type(1)", "el => el.innerHTML")
        # DOM에서 직접 라벨-값을 추출(구조 인식)
        pairs = page.evaluate(
            """
            () => {
              const root = document.querySelector('ul#docSummaryInfo li:nth-of-type(1)');
              const out = [];
              if (!root) return out;
              const push = (k, v) => { if (k && v) out.push({ label: k.trim(), value: v.trim() }); };

              // 1) dl/dt/dd
              root.querySelectorAll('dt').forEach(dt => {
                const dd = dt.nextElementSibling && dt.nextElementSibling.tagName.toLowerCase() === 'dd' ? dt.nextElementSibling : null;
                const k = dt.textContent || '';
                const v = dd ? dd.textContent : '';
                push(k, v);
              });

              // 2) table tr (th/td)
              root.querySelectorAll('tr').forEach(tr => {
                const cells = Array.from(tr.querySelectorAll('th,td')).map(c => (c.textContent || '').replace(/\s+/g,' ').trim()).filter(Boolean);
                if (cells.length >= 2) {
                  push(cells[0], cells.slice(1).join(' | '));
                }
              });

              // 3) .tit/.title/.key + next .cont/.val/.value/.data
              root.querySelectorAll('.tit, .title, .key').forEach(t => {
                const k = (t.textContent || '').trim();
                let sib = t.nextElementSibling;
                let v = '';
                if (sib && /cont|val|value|data/.test(sib.className)) v = sib.textContent || '';
                if (!v && sib && sib.querySelector) {
                  const cand = sib.querySelector('.cont, .val, .value, .data');
                  if (cand) v = cand.textContent || '';
                }
                if (k && v) push(k, v);
              });

              // 4) 일반 행에서 텍스트 스플릿(:, →, ->, -->)
              Array.from(root.children || []).forEach(row => {
                const t = (row.textContent || '').replace(/\s+/g,' ').trim();
                const m = t.match(/^(.{1,40}?)(?:\s*(?:[:：]|→|-->|->)\s+)(.+)$/);
                if (m) push(m[1], m[2]);
              });

              // 5) 중복 제거
              const seen = new Set();
              return out.filter(p => { const k = p.label + '|' + p.value; if (seen.has(k)) return false; seen.add(k); return true; });
            }
            """
        )

    if not pairs:
        # 텍스트 기반 재시도 (브라우저가 열려있는 동안 실행)
        try:
            li_text = page.inner_text("ul#docSummaryInfo li:nth-of-type(1)")  # type: ignore
        except Exception:
            li_text = ""
        if li_text:
            pairs = _extract_pairs_from_text_block(li_text)
    # 이제 브라우저를 닫는다
    try:
        browser.close()
    except Exception:
        pass
    structured = structure_docsummary_pairs(pairs)
    parsed = urlparse(target_url)
    qs = parse_qs(parsed.query)
    return {"url": target_url, "skey": qs.get("skey", [None])[0], "pairs": pairs, "structured": structured, "raw_html": li_html}


# ===== 단일 페이지: 엑셀 업로드 → 자동 파이프라인 → 미리보기/다운로드 =====
st.header("엑셀 업로드 → 자동 처리")

# session_state 초기화
if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False
if "chunks_records" not in st.session_state:
    st.session_state.chunks_records = []
if "uploaded_file_id" not in st.session_state:
    st.session_state.uploaded_file_id = None

uploaded = st.file_uploader("특허 엑셀 파일(.xlsx) 업로드", type=["xlsx"], accept_multiple_files=False, key="single_flow_uploader")

# 새로운 파일이 업로드되면 상태 초기화
if uploaded is not None:
    current_file_id = uploaded.file_id
    if st.session_state.uploaded_file_id != current_file_id:
        st.session_state.processing_complete = False
        st.session_state.chunks_records = []
        st.session_state.uploaded_file_id = current_file_id

# 이미 처리가 완료된 경우 결과만 표시
if uploaded is not None and st.session_state.processing_complete:
    st.success("✅ 처리가 완료되었습니다!")
    st.divider()

    chunks_records = st.session_state.chunks_records

    st.subheader(f"청크 미리보기 (총 {len(chunks_records)}개 중 최대 10개)")

    if chunks_records:
        preview_count = min(10, len(chunks_records))
        for i in range(preview_count):
            chunk_id = chunks_records[i].get('chunk_id', 'N/A')
            doc_id = chunks_records[i].get('doc_id', 'N/A')
            title = chunks_records[i].get('metadata', {}).get('title', '')
            text_len = len(chunks_records[i].get('text', ''))
            chunk_idx = chunks_records[i].get('chunk_index', '?')
            chunk_total = chunks_records[i].get('chunk_total', '?')
            para_range = chunks_records[i].get('paragraph_range', '')

            label = f"{i+1}. {chunk_id}"
            if para_range:
                label += f" | 문단 {para_range}"
            if title:
                label += f" | {title[:30]}{'...' if len(title) > 30 else ''}"
            label += f" | {text_len:,}자"
            with st.expander(label, expanded=(i==0)):
                st.json(chunks_records[i], expanded=False)

    # JSON 배열(샤딩) 다운로드
    def _make_json_shards(records: list[dict], max_bytes: int = 10 * 1024 * 1024) -> list[str]:
        import json as _json
        shards: list[str] = []
        current: list[dict] = []
        current_bytes = 2
        for rec in records:
            try:
                rec_str = _json.dumps(rec, ensure_ascii=False)
            except Exception:
                continue
            rec_size = len(rec_str.encode("utf-8"))
            sep = 1 if current else 0
            if current and (current_bytes + rec_size + sep) > max_bytes:
                shards.append(_json.dumps(current, ensure_ascii=False))
                current = []
                current_bytes = 2
                sep = 0
            current.append(rec)
            current_bytes += rec_size + sep
        if current:
            shards.append(_json.dumps(current, ensure_ascii=False))
        return shards

    chunks_shards = _make_json_shards(chunks_records)

    if len(chunks_shards) == 1:
        st.download_button(
            label=f"📥 chunks.json 다운로드 ({len(chunks_records)}개 청크)",
            data=chunks_shards[0].encode("utf-8"),
            file_name="patents.chunks.json",
            mime="application/json",
            type="primary",
        )
    else:
        st.info(f"💡 결과가 10MB를 초과하여 {len(chunks_shards)}개 파일로 분할되었습니다.")

        # ZIP 압축 다운로드 버튼
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i, shard in enumerate(chunks_shards, start=1):
                file_name = f"patents.chunks.part-{i:03d}.json"
                zip_file.writestr(file_name, shard.encode("utf-8"))

        zip_buffer.seek(0)
        st.download_button(
            label=f"📦 전체 다운로드 (ZIP, {len(chunks_shards)}개 파일)",
            data=zip_buffer.getvalue(),
            file_name="patents_chunks_all.zip",
            mime="application/zip",
            type="primary",
        )

        st.divider()
        st.caption("개별 파일 다운로드")

        # 개별 파일 다운로드 버튼
        cols = st.columns(min(3, len(chunks_shards)))
        for i, shard in enumerate(chunks_shards, start=1):
            col_idx = (i - 1) % 3
            with cols[col_idx]:
                st.download_button(
                    label=f"📥 Part {i}/{len(chunks_shards)}",
                    data=shard.encode("utf-8"),
                    file_name=f"patents.chunks.part-{i:03d}.json",
                    mime="application/json",
                    key=f"download_part_{i}",
                )

    st.stop()

# 처리가 필요한 경우 크롤링 실행
if uploaded is not None and not st.session_state.processing_complete:
    progress = st.progress(0)
    with st.status("파이프라인 시작", expanded=True) as status:
        try:
            # 1) 파일 로드
            st.write("1) 파일 로드")
            data = uploaded.read()
            progress.progress(5)

            # 2) 엑셀 파싱 및 메타데이터 수집
            st.write("2) 엑셀 파싱 및 메타데이터 수집")
            wb = load_workbook(filename=BytesIO(data), data_only=False, read_only=False)
            ws = wb.active
            progress.progress(20)

            # '출원번호' 헤더 탐색(링크 보관 열)
            target_col_idx = None
            header_row_idx = None
            matched_header_value = None
            search_rows = min(ws.max_row, 50)
            candidates = []
            for r in range(1, search_rows + 1):
                row_values = [normalize_header(c.value) for c in ws[r]]
                for c_idx, v in enumerate(row_values, start=1):
                    if not v:
                        continue
                    if v == "출원번호":
                        target_col_idx = c_idx
                        header_row_idx = r
                        matched_header_value = v
                        break
                    if "출원번호" in v:
                        candidates.append((c_idx, r, v))
                if target_col_idx is not None:
                    break
            if target_col_idx is None and candidates:
                target_col_idx, header_row_idx, matched_header_value = candidates[0]
            if target_col_idx is None:
                raise RuntimeError("'출원번호' 헤더를 찾지 못했습니다. 엑셀의 링크 열을 '출원번호'로 지정해주세요.")

            # 헤더 맵 구성
            headers = {}
            for col_idx, cell in enumerate(ws[header_row_idx], start=1):
                h = normalize_header(cell.value)
                if h:
                    headers[col_idx] = h

            # 행→URL 매핑
            row_to_url = {}
            for r in range(header_row_idx + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=target_col_idx)
                url, source = extract_url_from_cell(cell)
                if url and "sd.wips.co.kr" in url:
                    row_to_url[r] = url.strip()

            # 3) 원문 수집 및 RAG 청크 생성
            st.write("3) 특허 원문 수집 및 RAG 청크 생성")
            progress.progress(40)
            total_rows = max(ws.max_row - header_row_idx, 1)
            step_bar = st.progress(0, text="수집 진행률")

            def _lang_to_bcp47_inline(lang: str) -> str:
                if not lang:
                    return "und"
                code = lang.upper()
                if code in {"KR", "KO"}:
                    return "ko"
                if code in {"US", "EN"}:
                    return "en"
                if code in {"JP", "JA"}:
                    return "ja"
                if code in {"CN", "ZH"}:
                    return "zh"
                return code.lower()

            def _make_doc_id_inline(ctry: str | None, publication_number: str | None, application_number: str | None, fallback: str) -> str:
                c = (ctry or "").strip().upper()
                pn = (publication_number or "").strip()
                an = (application_number or "").strip()
                if c and pn:
                    return f"{c}:{pn}"
                if c and an:
                    return f"{c}:{an}"
                return fallback

            chunks_records: list[dict] = []

            for idx, r in enumerate(range(header_row_idx + 1, ws.max_row + 1), start=1):
                row_data = {}
                for col_idx, header_name in headers.items():
                    cell = ws.cell(row=r, column=col_idx)
                    value = cell.value
                    if header_name in ["출원일", "공개일", "등록일"]:
                        value = _normalize_date(str(value)) if value else ""
                    elif value is not None:
                        value = str(value).strip()
                    else:
                        value = ""
                    row_data[header_name] = value

                if not any(v for v in row_data.values()):
                    step_bar.progress(min(int(100 * idx / total_rows), 100))
                    continue

                url = row_to_url.get(r)
                selected_text = ""
                selected_lang = ""
                ctry = ""
                publication_number = row_data.get("공개번호", "")
                application_number = row_data.get("출원번호", "")
                title = row_data.get("발명의 명칭", "")

                if url:
                    try:
                        desc_out = fetch_wips_description(url)
                        by_lang = desc_out.get("descriptions_by_lang", {})
                        selected_text, selected_lang = _select_preferred_lang_text(by_lang)
                        ctry = (desc_out.get("doc", {}) or {}).get("ctry", "")
                    except Exception:
                        selected_text, selected_lang = "", ""

                doc_id_raw = application_number or publication_number or f"row_{r}"
                doc_id = _make_doc_id_inline(ctry, publication_number, application_number, doc_id_raw)
                processed_text = _decode_entities_and_normalize(selected_text)

                # 청크(chunks.json)
                if processed_text:
                    base_meta_for_chunk = {
                        "jurisdiction": (ctry or "").upper() or None,
                        "publication_number": publication_number,
                        "application_number": application_number,
                        "registration_number": row_data.get("등록번호", ""),
                        "filing_date": row_data.get("출원일", ""),
                        "publication_date": row_data.get("공개일", ""),
                        "registration_date": row_data.get("등록일", ""),
                        "assignees": [row_data.get("출원인", "")] if row_data.get("출원인") else [],
                        "title": title,
                        "legal_status": row_data.get("상태정보", ""),
                        "wips_url": url or "",
                    }
                    chunks = _build_chunk_records_for_doc(doc_id, base_meta_for_chunk, processed_text, selected_lang)
                    # 스키마 정규화
                    for ch in chunks:
                        if "id" in ch:
                            ch["chunk_id"] = ch.pop("id")
                        meta = ch.get("metadata", {}) or {}
                        lang_raw = meta.pop("lang", "")
                        ch["language"] = _lang_to_bcp47_inline(lang_raw)
                        if "chunk_index" in meta:
                            ch["chunk_index"] = meta.pop("chunk_index")
                        if "chunk_total" in meta:
                            ch["chunk_total"] = meta.pop("chunk_total")
                        if "paragraph_range" in meta:
                            ch["paragraph_range"] = meta.pop("paragraph_range")
                        ch["section"] = "description"
                        try:
                            ch["token_count"] = len(re.findall(r"\w+", ch.get("text", "")))
                        except Exception:
                            ch["token_count"] = 0
                    chunks_records.extend(chunks)

                step_bar.progress(min(int(100 * idx / total_rows), 100))

            # 결과를 session_state에 저장
            st.session_state.chunks_records = chunks_records
            st.session_state.processing_complete = True

            progress.progress(100)
            status.update(label="처리 완료", state="complete")
            st.success("✅ 크롤링 완료! 페이지가 새로고침됩니다.")
            st.rerun()
        except Exception as e:
            status.update(label="오류 발생", state="error")
            st.error(f"처리 중 오류가 발생했습니다: {e}")

    # 단일 플로우만 노출하고 기존 탭 UI는 렌더하지 않음
    st.stop()
